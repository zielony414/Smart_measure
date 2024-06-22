import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference
import xlwings as xw
import os
import numpy as np
import shutil

def import_data_to_sheet(ws, idx, path):
    headers = ['I_in [mA]', 'U_in [V]', 'I_out [mA]', 'U_out [V]', 'P_in [mW]', 'P_out [mV]', 'n[%]']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=8, column=col_idx, value=header)

    file_names = [
        f'inlet_amm{idx}.txt',
        f'inlet_volt{idx}.txt',
        f'out_amm{idx}.txt',
        f'out_volt{idx}.txt'
    ]

    for col_idx, file_name in enumerate(file_names, start=1):
        with open(path + '/' + file_name, 'r') as file:
            lines = file.readlines()
            for row_idx, line in enumerate(lines, start=9):  # Start from row 9
                data = line.split()
                if len(data) >= 2:
                    value = float(data[1])
                    ws.cell(row=row_idx, column=col_idx, value=value)

def calculate_and_create_chart(ws):
    for row in range(9, ws.max_row + 1):  # Start from row 9
        ws[f'E{row}'] = f"=A{row}*B{row}"
        ws[f'F{row}'] = f"=C{row}*D{row}"
        ws[f'G{row}'] = f"=F{row}/E{row}*100"

    chart = LineChart()
    chart.title = "Dependence of efficiency on input current"
    chart.x_axis.title = "I_in [mA]"
    chart.y_axis.title = "n [%]"

    x_data = Reference(ws, min_col=1, min_row=9, max_row=ws.max_row)  # Start from row 9
    y_data = Reference(ws, min_col=7, min_row=9, max_row=ws.max_row)

    chart.add_data(y_data, titles_from_data=True)
    chart.set_categories(x_data)

    chart.legend = None

    ws.add_chart(chart, "I10")

def data_analized(steps, path):

    #new file name
    final_filename = path + '/' + 'wyniki_badan.xlsx'
    #source file name
    source_filename = 'Template.xlsx'

    #copy source file to new file
    shutil.copy(source_filename, final_filename)

    #load new file
    new_wb = load_workbook(final_filename)

    #copy template sheet
    template_ws = new_wb['Data']
    for idx in range(steps):
        new_ws = new_wb.copy_worksheet(template_ws)
        new_ws.title = f"Data_Set_{idx}"

    ws = new_wb['Data']
    ws.title = f"combined_chart"


    #variables for combined chart
    all_I_in = []
    all_n = []
    all_u_in_means = []

    # import and calculate data
    for idx in range(steps):
        ws = new_wb[f"Data_Set_{idx}"]
        import_data_to_sheet(ws, idx, path)
        calculate_and_create_chart(ws)


    #save data to temp file for xlwing to open and calculate
    temp_filename = path + '/' + 'temp_dane_z_plikow_w_excelu_temp.xlsx'

    if os.path.exists(temp_filename):
        os.remove(temp_filename)

    new_wb.save(temp_filename)

    #calculate data
    app = xw.App(visible=False)
    book = xw.Book(temp_filename)

    for idx in range(4, len(book.sheets)):
        sheet = book.sheets[idx]
        sheet.api.Calculate()

    for idx in range(4, len(book.sheets)):
        sheet = book.sheets[idx]
        I_in = sheet.range('A9:A' + str(sheet.cells.last_cell.row)).value  # Start from row 9
        n = sheet.range('G9:G' + str(sheet.cells.last_cell.row)).value  # Start from row 9
        U_in = sheet.range('B9:B' + str(sheet.cells.last_cell.row)).value  # Start from row 9

        I_in = [i for i in I_in if i is not None]
        n = [i for i in n if i is not None]
        U_in = [u for u in U_in if u is not None]

        U_in = [float(u) for u in U_in if isinstance(u, (int, float, str)) and str(u).replace('.', '', 1).isdigit()]

        min_len = min(len(I_in), len(n))  # Ensure both lists have the same length
        I_in = I_in[:min_len]
        n = n[:min_len]

        all_I_in.append(I_in)
        all_n.append(n)

        if U_in:
            u_in_mean = np.mean(U_in)
        else:
            u_in_mean = 0

        all_u_in_means.append(u_in_mean)

    book.save()
    book.close()
    app.quit()

    # create combined chart
    colors = plt.get_cmap('tab10')  # colors

    plt.figure(figsize=(15, 10))  # chart size
    for idx in range(8):
        plt.plot(all_I_in[idx], all_n[idx], label=f'U_in ≈ {round(all_u_in_means[idx])}V', color=colors(idx / 7))

    plt.title('Dependence od efficiency on input current at different input voltage values')
    plt.xlabel('I_in [mA]')
    plt.ylabel('n [%]')
    plt.legend()
    plt.grid(True)

    # setting scale on chart
    plt.gca().xaxis.set_major_locator(plt.MultipleLocator(0.1))
    plt.gca().xaxis.set_minor_locator(plt.MultipleLocator(0.01))
    plt.gca().xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:.1f} mA'))
    plt.gca().yaxis.set_major_locator(plt.MultipleLocator(10))
    plt.gca().yaxis.set_minor_locator(plt.MultipleLocator(1))
    plt.gca().yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f'{y:.0f}%'))

    plt.gca().tick_params(axis='x', which='major', length=10, width=2)
    plt.gca().tick_params(axis='x', which='minor', length=5, width=1)

    plt.gca().xaxis.grid(True, which='minor')
    plt.gca().yaxis.grid(True, which='minor')

    # save image as PNG file
    plt.savefig(path + '/' + 'combined_chart.png')
    plt.close()

    # import chart to excel
    wb = load_workbook(temp_filename)
    combined_ws = wb[f"combined_chart"]
    img = Image(path + '/' + 'combined_chart.png')
    img.width, img.height = img.width * 1.5, img.height * 1.5
    combined_ws.add_image(img, 'A8')

    #save changes
    wb.save(final_filename)

    # remove temp file
    os.remove(temp_filename)

